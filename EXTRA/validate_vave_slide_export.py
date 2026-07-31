"""Generate and validate the current VAVE slide-pack workbook."""
from pathlib import Path
import os
import sys

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
WORKSPACE_DIR = APP_DIR.parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from utils.excel_export import export_projects  # noqa: E402
from utils.vpmt_io import load_projects  # noqa: E402


SOURCE_FILE = APP_DIR / "DG-PH2-6-25.vpmt"
OUTPUT_DIR = WORKSPACE_DIR / "outputs" / "vave-slide-pack"
OUTPUT_FILE = Path(
    os.environ.get(
        "VPM_VAVE_VALIDATION_OUTPUT",
        str(OUTPUT_DIR / "VAVE-Slide-Pack-Validation.xlsx"),
    )
).resolve()


def main():
    if not SOURCE_FILE.exists():
        raise FileNotFoundError(f"Source project was not found: {SOURCE_FILE}")

    projects = load_projects(str(SOURCE_FILE))
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    export_projects(projects, str(OUTPUT_FILE))

    workbook = load_workbook(OUTPUT_FILE, data_only=False)
    data_sheets = [name for name in workbook.sheetnames if name.endswith(" VAVE Data")]
    if not data_sheets:
        raise AssertionError("No VAVE Data sheet was generated.")

    preview_sheets = [
        name for name in workbook.sheetnames
        if name not in data_sheets
        and not name.endswith(" Tasks")
        and not name.endswith(" Metadata")
    ]
    if not preview_sheets:
        raise AssertionError("No VAVE preview sheets were generated.")

    for sheet_name in data_sheets:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if any(row[13] == "In Progress" for row in rows):
            raise AssertionError(f"{sheet_name} still contains slide-facing 'In Progress'.")
        groups = {}
        for row in rows:
            if row[14] == 1:
                groups.setdefault(row[2], 0.0)
                groups[row[2]] += float(row[10] or 0)
        print(f"{sheet_name}: {len(rows)} slide rows")
        for title, total in groups.items():
            print(f"  {title}: ${total:,.2f}")

    for sheet_name in preview_sheets:
        sheet = workbook[sheet_name]
        if any(cell.value == "In Progress" for row in sheet.iter_rows() for cell in row):
            raise AssertionError(f"{sheet_name} still contains 'In Progress'.")
        print(f"Preview OK: {sheet_name}")

    print(f"\nValidation workbook created:\n{OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"\nVALIDATION FAILED: {exc}")
        raise SystemExit(1)
