"""
Excel (.xlsx) export for one or many projects.

Layout per project (two sheets):
  "<Project> Tasks"   — flat rows: Level, Task, Start, End, Duration,
                        Status, Owner, Depends On, Notes
  "<Project> Metadata" — owners, holidays, exclude_weekends

Dates are written as ISO strings (YYYY-MM-DD) to sidestep Excel's
timezone-serial quirks. Hierarchy is encoded as a numeric "Level"
column (0 = root) rather than merged cells — safer for filtering
and sorting inside Excel.

Requires `openpyxl`. Kept as an optional dependency: the import is
deferred so the main app still runs if the package isn't installed.
"""
from typing import List, Dict
from models.task_node import TaskNode


TASK_COLUMNS = [
    "Level", "Task Name", "Start", "End", "Duration",
    "Status", "Owner", "Depends On", "Notes",
]


def _safe_sheet_name(name: str, suffix: str, seen: set) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\, unique per workbook."""
    base = (name or "Project").strip() or "Project"
    for ch in "[]:*?/\\":
        base = base.replace(ch, "_")
    # Leave room for the suffix
    room = 31 - len(suffix) - 1
    if room < 3:
        room = 3
    base = base[:room]
    candidate = f"{base} {suffix}"
    # Deduplicate if two projects share a name
    n = 2
    original = candidate
    while candidate in seen:
        tail = f" ({n})"
        candidate = (original[: 31 - len(tail)] + tail)
        n += 1
    seen.add(candidate)
    return candidate


def _flatten(roots: List[TaskNode]) -> List[Dict]:
    """Post-order-preserving depth-first flatten. Matches tree visual order."""
    rows = []

    def visit(node: TaskNode, depth: int):
        pred_id = node.predecessor_id or ""
        rows.append({
            "Level": depth,
            "Task Name": node.name or "",
            "Start": node.start_date or "",
            "End": node.end_date or "",
            "Duration": node.duration or "",
            "Status": node.status or "",
            "Owner": node.owner or "",
            "Depends On": pred_id,  # resolved to name by caller
            "Notes": (node.notes or "").replace("\r", ""),
        })
        for c in node.children:
            visit(c, depth + 1)

    for r in roots:
        visit(r, 0)
    return rows


def _resolve_pred_names(rows: List[Dict], roots: List[TaskNode]):
    """Replace Depends On ids with the target task name (in place)."""
    id_to_name = {}

    def collect(n: TaskNode):
        id_to_name[n.id] = n.name
        for c in n.children:
            collect(c)
    for r in roots:
        collect(r)

    for row in rows:
        pid = row["Depends On"]
        if pid:
            row["Depends On"] = id_to_name.get(pid, "(missing)")


def export_projects(projects: List[Dict], filename: str):
    """
    Write every project to a single workbook.

    `projects` is a list of dicts like:
        {"name": str, "metadata": dict, "roots": [TaskNode, ...]}
    — the same shape that vpmt_io.load_projects returns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    # Workbook starts with one default sheet — remove it, we'll add our own.
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    wrap = Alignment(wrap_text=True, vertical="top")
    used_names = set()

    for proj in projects:
        name = proj.get("name") or "Project"
        roots = proj.get("roots", [])
        metadata = proj.get("metadata", {}) or {}

        # --- Tasks sheet ---
        ws = wb.create_sheet(_safe_sheet_name(name, "Tasks", used_names))
        ws.append(TASK_COLUMNS)
        for col_idx in range(1, len(TASK_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        rows = _flatten(roots)
        _resolve_pred_names(rows, roots)
        for row in rows:
            ws.append([row[c] for c in TASK_COLUMNS])
            # Indent the Task Name cell to mirror the outline
            last_row = ws.max_row
            tn_cell = ws.cell(row=last_row, column=2)
            tn_cell.alignment = Alignment(indent=int(row["Level"]), wrap_text=True, vertical="top")
            notes_cell = ws.cell(row=last_row, column=TASK_COLUMNS.index("Notes") + 1)
            notes_cell.alignment = wrap

        # Set reasonable column widths
        widths = {
            "Level": 8, "Task Name": 42, "Start": 12, "End": 12,
            "Duration": 10, "Status": 14, "Owner": 18,
            "Depends On": 28, "Notes": 60,
        }
        for i, name_col in enumerate(TASK_COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(name_col, 14)
        ws.freeze_panes = "A2"

        # --- Metadata sheet ---
        ms = wb.create_sheet(_safe_sheet_name(name, "Metadata", used_names))
        ms.append(["Field", "Value"])
        ms.cell(row=1, column=1).font = header_font
        ms.cell(row=1, column=1).fill = header_fill
        ms.cell(row=1, column=2).font = header_font
        ms.cell(row=1, column=2).fill = header_fill

        owners = metadata.get("owners", []) or []
        holidays = metadata.get("holidays", []) or []
        excl = metadata.get("exclude_weekends", True)
        ms.append(["Owners", ", ".join(owners)])
        ms.append(["Holidays", ", ".join(holidays)])
        ms.append(["Exclude Weekends", "Yes" if excl else "No"])
        ms.column_dimensions["A"].width = 20
        ms.column_dimensions["B"].width = 80

    if not wb.sheetnames:
        # Pathological: zero projects — leave a stub so the file opens cleanly.
        wb.create_sheet("Empty")

    wb.save(filename)
