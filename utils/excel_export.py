"""
Excel (.xlsx) export for one or many projects.

Base layout per project (two sheets):
  "<Project> Tasks"   — flat rows: Level, Task, Start, End, Duration,
                        Status, Owner, Depends On, Notes
  "<Project> Metadata" — owners, holidays, exclude_weekends

Dates are written as ISO strings (YYYY-MM-DD) to sidestep Excel's
timezone-serial quirks. Hierarchy is encoded as a numeric "Level"
column (0 = root) rather than merged cells — safer for filtering
and sorting inside Excel.

VAVE projects also receive a machine-readable ``VAVE Data`` sheet and one
slide-ready preview sheet per VAVE activity group.  The preview pagination is
deliberately deterministic so a later PowerPoint generator does not have to
infer layout:

* 1-7 rows: one full-width table
* 8-14 rows: two balanced columns
* 15+ rows: continuation sheets, at most 14 rows per sheet

Requires `openpyxl`. Kept as an optional dependency: the import is
deferred so the main app still runs if the package isn't installed.
"""
from datetime import datetime
from math import ceil
from typing import List, Dict
from models.task_node import TaskNode


TASK_COLUMNS = [
    "Level", "Task Name", "Start", "End", "Duration",
    "Status", "Owner", "Depends On", "Notes",
]

VAVE_TASK_COLUMNS = ["Potential $", "Realized $"]

VAVE_DATA_COLUMNS = [
    "Project", "Slide Order", "Slide Title", "Page", "Column",
    "Row On Page", "Task Order", "Task Name", "Potential $", "Realized $",
    "Savings", "Savings Basis", "End Date", "Status", "Include In Total",
    "Source Task ID",
]

VAVE_ROWS_PER_COLUMN = 7
VAVE_ROWS_PER_PAGE = 14

_NAVY = "245780"
_BLUE = "4472C4"
_ORANGE = "F47C20"
_GREEN = "70AD47"
_GRAY = "A5A5A5"
_HEADER_GRAY = "E7EDF5"
_ROW_GRAY = "F1F4F8"
_WHITE = "FFFFFF"
_TEXT = "202020"


def _safe_sheet_name(name: str, suffix: str, seen: set) -> str:
    """Excel sheet names: max 31 chars, no []:*?/\\, unique per workbook."""
    base = (name or "Project").strip() or "Project"
    for ch in "[]:*?/\\":
        base = base.replace(ch, "_")
    # Leave room for the suffix
    room = 31 - len(suffix) - 1
    if room < 3:
        room = 3
    base = base[:room]
    candidate = f"{base} {suffix}"
    # Deduplicate if two projects share a name
    n = 2
    original = candidate
    while candidate in seen:
        tail = f" ({n})"
        candidate = (original[: 31 - len(tail)] + tail)
        n += 1
    seen.add(candidate)
    return candidate


def _unique_sheet_name(name: str, seen: set) -> str:
    """Return a valid, unique Excel sheet name without forcing a suffix."""
    base = (name or "Sheet").strip() or "Sheet"
    for ch in "[]:*?/\\":
        base = base.replace(ch, "_")
    base = base[:31]
    candidate = base
    n = 2
    while candidate in seen:
        tail = f" ({n})"
        candidate = base[:31 - len(tail)] + tail
        n += 1
    seen.add(candidate)
    return candidate


def _flatten(roots: List[TaskNode]) -> List[Dict]:
    """Post-order-preserving depth-first flatten. Matches tree visual order."""
    rows = []

    def visit(node: TaskNode, depth: int):
        pred_id = node.predecessor_id or ""
        owner_text = node.owner or ""
        if getattr(node, "waiting_on", ""):
            owner_text = f"Waiting on {node.waiting_on}"
            if getattr(node, "waiting_since", None):
                owner_text += f" since {node.waiting_since}"
            if node.owner:
                owner_text += f" (owner: {node.owner})"
        rows.append({
            "Level": depth,
            "Task Name": node.name or "",
            "Start": node.start_date or "",
            "End": node.end_date or "",
            "Duration": node.duration or "",
            "Potential $": node.vave_display_potential(),
            "Realized $": node.vave_display_realized(),
            "Status": node.status or "",
            "Owner": owner_text,
            "Depends On": pred_id,  # resolved to name by caller
            "Notes": (node.notes or "").replace("\r", ""),
        })
        for c in node.children:
            visit(c, depth + 1)

    for r in roots:
        visit(r, 0)
    return rows


def _resolve_pred_names(rows: List[Dict], roots: List[TaskNode]):
    """Replace Depends On ids with the target task name (in place)."""
    id_to_name = {}

    def collect(n: TaskNode):
        id_to_name[n.id] = n.name
        for c in n.children:
            collect(c)
    for r in roots:
        collect(r)

    for row in rows:
        pid = row["Depends On"]
        if pid:
            row["Depends On"] = id_to_name.get(pid, "(missing)")


def _has_manual_vave_value(node: TaskNode) -> bool:
    """Only manual task dollars create slide rows; roll-up dollars do not."""
    return (getattr(node, "vave_potential", None) is not None or
            getattr(node, "vave_realized", None) is not None)


def _collect_vave_sections(roots: List[TaskNode]) -> List[Dict]:
    """Collect deepest useful VAVE groups in visual tree order.

    A section is a node with one or more immediate children carrying manual
    VAVE dollars.  This maps the current tracker hierarchy directly to the
    reference deck: e.g. ``Cassette VAVE activities`` becomes a slide title
    and its dollar-bearing children become the slide rows.
    """
    sections = []

    def visit(node: TaskNode):
        rows = [child for child in node.children if _has_manual_vave_value(child)]
        if rows:
            sections.append({"title": node.name or "VAVE activities", "nodes": rows})
        for child in node.children:
            visit(child)

    root_rows = [root for root in roots if _has_manual_vave_value(root)]
    if root_rows:
        sections.append({"title": "VAVE activities", "nodes": root_rows})
    for root in roots:
        visit(root)
    return sections


def _number_or_none(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _date_or_text(value):
    """Use typed Excel dates when the tracker value is a valid ISO date."""
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return value


def _vave_row(node: TaskNode) -> Dict:
    potential = _number_or_none(node.vave_display_potential())
    realized = _number_or_none(node.vave_display_realized())
    explicit_realized = _number_or_none(getattr(node, "vave_realized", None))

    if explicit_realized is not None:
        savings = explicit_realized
        savings_basis = "Realized"
        display_status = "REALIZED"
    else:
        savings = potential
        savings_basis = "Potential"
        raw_status = node.status or ""
        display_status = "On Track" if raw_status.strip().lower() == "in progress" else raw_status

    return {
        "node": node,
        "task_name": node.name or "",
        "potential": potential,
        "realized": realized,
        "savings": savings,
        "savings_basis": savings_basis,
        "end_date": _date_or_text(node.end_date),
        "status": display_status,
        # Numeric 1/0 is intentionally more portable than TRUE/FALSE across
        # Excel readers and formula engines used by downstream slide tooling.
        "include": 1,
    }


def _page_layout(row_count: int) -> List[Dict]:
    """Describe stable page/column positions for a section's rows."""
    layout = []
    for zero_index in range(row_count):
        page = zero_index // VAVE_ROWS_PER_PAGE + 1
        index_on_page = zero_index % VAVE_ROWS_PER_PAGE
        rows_on_page = min(VAVE_ROWS_PER_PAGE,
                           row_count - (page - 1) * VAVE_ROWS_PER_PAGE)
        if rows_on_page <= VAVE_ROWS_PER_COLUMN:
            column = "Full"
            row_on_page = index_on_page + 1
        else:
            left_count = ceil(rows_on_page / 2)
            if index_on_page < left_count:
                column = "Left"
                row_on_page = index_on_page + 1
            else:
                column = "Right"
                row_on_page = index_on_page - left_count + 1
        layout.append({
            "page": page,
            "column": column,
            "row_on_page": row_on_page,
        })
    return layout


def _style_status_cell(cell, status: str, Font, PatternFill, Alignment):
    normalized = (status or "").strip().lower()
    if normalized in ("completed", "realized"):
        color = _GREEN
    elif normalized in ("in progress", "on track"):
        color = _ORANGE
    else:
        color = _GRAY
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_vave_data_sheet(wb, project_name: str, sections: List[Dict],
                           used_names: set, Font, PatternFill, Alignment):
    """Write the stable handoff contract consumed by future slide tooling."""
    ws = wb.create_sheet(_safe_sheet_name(project_name, "VAVE Data", used_names))
    ws.sheet_view.showGridLines = False
    ws.append(VAVE_DATA_COLUMNS)

    for col_idx in range(1, len(VAVE_DATA_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Arial", size=10, bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    output_row = 2
    slide_order = 1
    for section in sections:
        rows = [_vave_row(node) for node in section["nodes"]]
        positions = _page_layout(len(rows))
        for task_order, (row, pos) in enumerate(zip(rows, positions), start=1):
            ws.append([
                project_name,
                slide_order + pos["page"] - 1,
                section["title"],
                pos["page"],
                pos["column"],
                pos["row_on_page"],
                task_order,
                row["task_name"],
                row["potential"],
                row["realized"],
                row["savings"],
                row["savings_basis"],
                row["end_date"],
                row["status"],
                row["include"],
                row["node"].id,
            ])
            output_row += 1
        slide_order += max(1, ceil(len(rows) / VAVE_ROWS_PER_PAGE))

    if ws.max_row > 1:
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor=_ROW_GRAY)
            ws.cell(row=row_idx, column=9).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=10).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=11).number_format = '$#,##0.00'
            ws.cell(row=row_idx, column=13).number_format = "m/d/yyyy"
            for cell in ws[row_idx]:
                cell.font = Font(name="Arial", size=10, color=_TEXT)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.auto_filter.ref = f"A1:P{ws.max_row}"

    widths = [20, 12, 34, 8, 10, 12, 11, 48, 13, 13, 13, 15, 13, 16, 16, 38]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    return ws


def _write_preview_block(ws, rows: List[Dict], start_col: int, header_row: int,
                         full_width: bool, Font, PatternFill, Alignment):
    """Write one visual table block."""
    if full_width:
        # A:F task name, G savings, H end, I status.
        task_start, task_end = start_col, start_col + 5
        savings_col, end_col, status_col = start_col + 6, start_col + 7, start_col + 8
    else:
        # A task name, B savings, C end, D status (or F:I on the right).
        task_start = task_end = start_col
        savings_col, end_col, status_col = start_col + 1, start_col + 2, start_col + 3

    if task_end > task_start:
        ws.merge_cells(start_row=header_row, start_column=task_start,
                       end_row=header_row, end_column=task_end)
    headers = [
        (task_start, "Task Name"),
        (savings_col, "Savings"),
        (end_col, "End"),
        (status_col, "Status"),
    ]
    for col, value in headers:
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.font = Font(name="Arial", size=10, bold=True, color=_TEXT)
        cell.fill = PatternFill("solid", fgColor=_HEADER_GRAY)
        cell.alignment = Alignment(vertical="center",
                                   horizontal="center" if col != task_start else "left")
    for col in range(task_start, status_col + 1):
        ws.cell(row=header_row, column=col).fill = PatternFill("solid", fgColor=_HEADER_GRAY)

    for offset, row in enumerate(rows, start=1):
        sheet_row = header_row + offset
        if task_end > task_start:
            ws.merge_cells(start_row=sheet_row, start_column=task_start,
                           end_row=sheet_row, end_column=task_end)
        if offset % 2 == 1:
            for col in range(task_start, status_col + 1):
                ws.cell(row=sheet_row, column=col).fill = PatternFill("solid", fgColor=_ROW_GRAY)

        task_cell = ws.cell(row=sheet_row, column=task_start, value=row["task_name"])
        task_cell.font = Font(name="Arial", size=10, color=_TEXT)
        task_cell.alignment = Alignment(vertical="center", wrap_text=True)

        savings_cell = ws.cell(row=sheet_row, column=savings_col, value=row["savings"])
        savings_cell.number_format = '$#,##0.00'
        savings_cell.font = Font(name="Arial", size=10, bold=True, color=_TEXT)
        savings_cell.alignment = Alignment(horizontal="right", vertical="center")

        end_cell = ws.cell(row=sheet_row, column=end_col, value=row["end_date"])
        end_cell.number_format = "m/d/yyyy"
        end_cell.font = Font(name="Arial", size=10, color=_TEXT)
        end_cell.alignment = Alignment(horizontal="center", vertical="center")

        status_cell = ws.cell(row=sheet_row, column=status_col, value=row["status"])
        _style_status_cell(status_cell, row["status"], Font, PatternFill, Alignment)
        ws.row_dimensions[sheet_row].height = 39


def _write_vave_preview_sheet(wb, project_name: str, section_title: str,
                              page_number: int, rows: List[Dict], color: str,
                              data_sheet_name: str, data_last_row: int,
                              used_names: set, Font, PatternFill, Alignment):
    suffix = "" if page_number == 1 else f" ({page_number})"
    ws = wb.create_sheet(_unique_sheet_name(section_title + suffix, used_names))
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    display_title = section_title if page_number == 1 else f"{section_title} (continued)"
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = display_title
    title_cell.fill = PatternFill("solid", fgColor=_NAVY)
    title_cell.font = Font(name="Arial", size=20, bold=True, color=_WHITE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A3:G3")
    section_cell = ws["A3"]
    section_cell.value = display_title
    section_cell.fill = PatternFill("solid", fgColor=color)
    section_cell.font = Font(name="Arial", size=11, bold=True, color=_WHITE)
    section_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["H3"] = "Total:"
    ws["H3"].fill = PatternFill("solid", fgColor=color)
    ws["H3"].font = Font(name="Arial", size=11, bold=True, color=_WHITE)
    ws["H3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["I3"].fill = PatternFill("solid", fgColor=color)
    ws["I3"].font = Font(name="Arial", size=11, bold=True, color=_WHITE)
    ws["I3"].alignment = Alignment(horizontal="right", vertical="center")
    ws["I3"].number_format = '$#,##0.00'
    ws.row_dimensions[3].height = 28

    if len(rows) <= VAVE_ROWS_PER_COLUMN:
        _write_preview_block(ws, rows, 1, 5, True, Font, PatternFill, Alignment)
    else:
        left_count = ceil(len(rows) / 2)
        _write_preview_block(
            ws, rows[:left_count], 1, 5, False, Font, PatternFill, Alignment)
        _write_preview_block(
            ws, rows[left_count:], 6, 5, False, Font, PatternFill, Alignment)

    # Every page for a section shows the section total, not merely the rows on
    # that page.  The formula intentionally reads the machine-contract sheet
    # and honors its Include In Total flag, making exclusions explicit and
    # auditable instead of allowing a manually typed banner total to drift.
    quoted_sheet = data_sheet_name.replace("'", "''")
    quoted_title = section_title.replace('"', '""')
    ws["I3"] = (
        f'=SUMIFS(\'{quoted_sheet}\'!$K$2:$K${data_last_row},'
        f'\'{quoted_sheet}\'!$C$2:$C${data_last_row},"{quoted_title}",'
        f'\'{quoted_sheet}\'!$O$2:$O${data_last_row},1)'
    )

    widths = {"A": 42, "B": 12, "C": 13, "D": 16, "E": 3,
              "F": 42, "G": 12, "H": 13, "I": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:I{max(12, 5 + min(len(rows), VAVE_ROWS_PER_COLUMN))}"
    ws.sheet_properties.tabColor = color
    ws.sheet_view.zoomScale = 85
    return ws


def _write_vave_slide_pack(wb, project_name: str, roots: List[TaskNode],
                           used_names: set, Font, PatternFill, Alignment):
    sections = _collect_vave_sections(roots)
    if not sections:
        return

    data_ws = _write_vave_data_sheet(
        wb, project_name, sections, used_names, Font, PatternFill, Alignment)

    colors = [_BLUE, _ORANGE]
    for section_index, section in enumerate(sections):
        rows = [_vave_row(node) for node in section["nodes"]]
        color = colors[section_index % len(colors)]
        for page_start in range(0, len(rows), VAVE_ROWS_PER_PAGE):
            page_rows = rows[page_start:page_start + VAVE_ROWS_PER_PAGE]
            page_number = page_start // VAVE_ROWS_PER_PAGE + 1
            _write_vave_preview_sheet(
                wb, project_name, section["title"], page_number, page_rows,
                color, data_ws.title, data_ws.max_row, used_names,
                Font, PatternFill, Alignment)


def export_projects(projects: List[Dict], filename: str):
    """
    Write every project to a single workbook.

    `projects` is a list of dicts like:
        {"name": str, "metadata": dict, "roots": [TaskNode, ...]}
    — the same shape that vpmt_io.load_projects returns.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    # Workbook starts with one default sheet — remove it, we'll add our own.
    default = wb.active
    wb.remove(default)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    wrap = Alignment(wrap_text=True, vertical="top")
    used_names = set()

    for proj in projects:
        name = proj.get("name") or "Project"
        roots = proj.get("roots", [])
        metadata = proj.get("metadata", {}) or {}
        task_columns = list(TASK_COLUMNS)
        if proj.get("is_vave"):
            insert_at = task_columns.index("Status")
            task_columns[insert_at:insert_at] = VAVE_TASK_COLUMNS

        # --- Tasks sheet ---
        ws = wb.create_sheet(_safe_sheet_name(name, "Tasks", used_names))
        ws.append(task_columns)
        for col_idx in range(1, len(task_columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        rows = _flatten(roots)
        _resolve_pred_names(rows, roots)
        for row in rows:
            ws.append([row[c] for c in task_columns])
            # Indent the Task Name cell to mirror the outline
            last_row = ws.max_row
            tn_cell = ws.cell(row=last_row, column=2)
            tn_cell.alignment = Alignment(indent=int(row["Level"]), wrap_text=True, vertical="top")
            notes_cell = ws.cell(row=last_row, column=task_columns.index("Notes") + 1)
            notes_cell.alignment = wrap
            for money_col in VAVE_TASK_COLUMNS:
                if money_col in task_columns:
                    cell = ws.cell(row=last_row, column=task_columns.index(money_col) + 1)
                    cell.number_format = '$#,##0.0'

        # Set reasonable column widths
        widths = {
            "Level": 8, "Task Name": 42, "Start": 12, "End": 12,
            "Duration": 10, "Potential $": 14, "Realized $": 14,
            "Status": 14, "Owner": 18,
            "Depends On": 28, "Notes": 60,
        }
        for i, name_col in enumerate(task_columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = widths.get(name_col, 14)
        ws.freeze_panes = "A2"

        # --- Metadata sheet ---
        ms = wb.create_sheet(_safe_sheet_name(name, "Metadata", used_names))
        ms.append(["Field", "Value"])
        ms.cell(row=1, column=1).font = header_font
        ms.cell(row=1, column=1).fill = header_fill
        ms.cell(row=1, column=2).font = header_font
        ms.cell(row=1, column=2).fill = header_fill

        owners = metadata.get("owners", []) or []
        holidays = metadata.get("holidays", []) or []
        excl = metadata.get("exclude_weekends", True)
        ms.append(["Owners", ", ".join(owners)])
        ms.append(["Holidays", ", ".join(holidays)])
        ms.append(["Exclude Weekends", "Yes" if excl else "No"])
        ms.column_dimensions["A"].width = 20
        ms.column_dimensions["B"].width = 80

        if proj.get("is_vave"):
            _write_vave_slide_pack(
                wb, name, roots, used_names, Font, PatternFill, Alignment)

    if not wb.sheetnames:
        # Pathological: zero projects — leave a stub so the file opens cleanly.
        wb.create_sheet("Empty")

    wb.save(filename)
