"""
build_eol_workbook.py  v3 — Formula-Driven
Every calculated cell is an Excel formula referencing RawData.
Python only writes: source data rows, formula strings, formatting, and charts.
"""
import os
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, "SAVE FILES", "Run test data start to 4-27-26.xlsx")
OUT_DIR   = os.path.join(BASE, "eol_output")
OUT_FILE  = os.path.join(OUT_DIR, "EOL_BellCurves_v3d.xlsx")

BOM_MAP = {"CRS0000675": "LT", "CRS0000674": "MT"}

# (raw_col_name, label, RawData_column_letter)
MEAS_COLS = [
    ("RunTestSuctionTubeTemp",            "Suction Tube Temp",       "G"),
    ("RunTestEvapInletRetBendATemp",      "Evap Inlet Ret Bend A",   "H"),
    ("RunTestEvapInletRetBendBTemp",      "Evap Inlet Ret Bend B",   "I"),
    ("RunTestEvapOutletAirTemp",          "Evap Outlet Air Temp",    "J"),
    ("RunTestEvap_B_OutletAirTemp",       "Evap B Outlet Air Temp",  "K"),
    ("RunTestInletOutletAirTempDelta",    "Inlet/Outlet Delta",      "L"),
    ("RunTestInletOutlet_B_AirTempDelta", "Inlet/Outlet B Delta",    "M"),
    ("RunTestAmbientTemperature",         "Ambient Temp",            "N"),
    ("RunTestVoltageStartup",             "Startup Voltage",         "O"),
    ("RunTestFinalVoltage",               "Final Voltage",           "P"),
    ("RunTestCurrentStartup",             "Startup Current",         "Q"),
    ("RunTestFinalCurrent",               "Final Current",           "R"),
]
N_BINS = 30
LAST_ROW = 941  # row 1 = header, rows 2-941 = 940 data rows

# ── Styles ─────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill("solid", fgColor="EEF4FB")
RED_FILL = PatternFill("solid", fgColor="FF9999")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
YLW_FILL = PatternFill("solid", fgColor="FFEB9C")

def hdr_cell(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.fill, c.font = HDR_FILL, HDR_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    return c

# ── Load data ───────────────────────────────────────────────────────────────
def load():
    df = pd.read_excel(DATA_FILE, engine="openpyxl")
    df["UnitType"] = df["BomNo"].map(BOM_MAP)
    df["TestDate"] = pd.to_datetime(df["TestDate"]).dt.date
    id_cols = ["SerialNo", "BomNo", "UnitType", "ModelNo", "TestDate", "OverallPassFail"]
    meas_present = [c for c, _, _ in MEAS_COLS if c in df.columns]
    return df[id_cols + meas_present].copy(), meas_present

# ── Sheet 1: RawData ────────────────────────────────────────────────────────
def build_rawdata(wb, df, meas_present):
    ws = wb.create_sheet("RawData")
    id_cols = ["SerialNo", "BomNo", "UnitType", "ModelNo", "TestDate", "OverallPassFail"]
    z_headers = [f"Z_{c}" for c in meas_present]
    headers = id_cols + meas_present + z_headers

    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = max(14, len(h) * 0.85)

    n_id = len(id_cols)
    n_meas = len(meas_present)
    z_col_start = n_id + n_meas + 1

    # Write source data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(id_cols, 1):
            ws.cell(row=ri, column=ci, value=row[col])
        for mi, col in enumerate(meas_present):
            ws.cell(row=ri, column=n_id + mi + 1, value=row.get(col))

    # Z-score formulas referencing Config sheet (mean in col F, stdev in col G)
    # Config rows: LT meas 0-11 = rows 2-13, MT meas 0-11 = rows 14-25
    for ri in range(2, LAST_ROW + 1):
        for mi in range(n_meas):
            val_cell = f"{get_column_letter(n_id + mi + 1)}{ri}"
            ut_cell = f"$C{ri}"
            # Lookup: if UnitType=LT use Config row (mi+2), if MT use Config row (mi+14)
            config_row_lt = mi + 2
            config_row_mt = mi + 2 + len(MEAS_COLS)
            mean_ref = f'IF({ut_cell}="LT",Config!$F${config_row_lt},Config!$F${config_row_mt})'
            sd_ref   = f'IF({ut_cell}="LT",Config!$G${config_row_lt},Config!$G${config_row_mt})'
            formula = f'=IF(OR({val_cell}=0,{val_cell}=""),"",ABS(({val_cell}-{mean_ref})/{sd_ref}))'
            ws.cell(row=ri, column=z_col_start + mi, value=formula)

    # Conditional formatting: red on z > 3
    for mi in range(n_meas):
        cl = get_column_letter(z_col_start + mi)
        ws.conditional_formatting.add(
            f"{cl}2:{cl}{LAST_ROW}",
            CellIsRule(operator="greaterThan", formula=["3"], fill=RED_FILL)
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{LAST_ROW}"
    print(f"  RawData: {LAST_ROW-1} rows, {len(headers)} cols")

# ── Sheet 2: Config ─────────────────────────────────────────────────────────
def build_config(wb):
    ws = wb.create_sheet("Config")
    headers = ["Unit", "Measurement", "RawDataCol", "n_all", "n_pass",
               "Mean_pass", "StdDev_pass", "Min_all", "Max_all",
               "BinWidth", "BinStart"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = 16

    rng = f"$2:${LAST_ROW}"  # data range in RawData
    ri = 2
    for ut in ["LT", "MT"]:
        for _, label, col_letter in MEAS_COLS:
            ws.cell(row=ri, column=1, value=ut)
            ws.cell(row=ri, column=2, value=label)
            ws.cell(row=ri, column=3, value=col_letter)

            val_rng = f"RawData!${col_letter}$2:${col_letter}${LAST_ROW}"
            ut_rng  = f"RawData!$C$2:$C${LAST_ROW}"
            pf_rng  = f"RawData!$F$2:$F${LAST_ROW}"

            # D: n_all (count non-zero for this unit)
            ws.cell(row=ri, column=4,
                    value=f'=COUNTIFS({ut_rng},"{ut}",{val_rng},"<>0")')

            # E: n_pass
            ws.cell(row=ri, column=5,
                    value=f'=COUNTIFS({ut_rng},"{ut}",{pf_rng},"Pass",{val_rng},"<>0")')

            # F: Mean_pass
            ws.cell(row=ri, column=6,
                    value=f'=IFERROR(AVERAGEIFS({val_rng},{ut_rng},"{ut}",{pf_rng},"Pass",{val_rng},"<>0"),0)')

            # G: StdDev_pass (SUMPRODUCT-based since no STDEVIFS)
            mean_ref = f"$F{ri}"
            npass_ref = f"$E{ri}"
            ws.cell(row=ri, column=7,
                    value=f'=IFERROR(SQRT(SUMPRODUCT(({ut_rng}="{ut}")*({pf_rng}="Pass")*({val_rng}<>0)*({val_rng}-{mean_ref})^2)/({npass_ref}-1)),0)')

            # H: Min_all — AGGREGATE(15,6,arr,1) = SMALL ignoring errors, k=1 = MIN
            # Division trick: values / condition → #DIV/0! for non-matches, AGGREGATE skips them
            agg_cond = f'(({ut_rng}="{ut}")*({val_rng}<>0))'
            ws.cell(row=ri, column=8,
                    value=f'=IFERROR(AGGREGATE(15,6,{val_rng}/{agg_cond},1),0)')

            # I: Max_all — AGGREGATE(14,6,arr,1) = LARGE ignoring errors, k=1 = MAX
            ws.cell(row=ri, column=9,
                    value=f'=IFERROR(AGGREGATE(14,6,{val_rng}/{agg_cond},1),0)')

            # J: BinWidth = (Max - Min + StdDev) / N_BINS
            ws.cell(row=ri, column=10,
                    value=f'=IFERROR(($I{ri}-$H{ri}+$G{ri})/{N_BINS},1)')

            # K: BinStart = Min - 0.5 * StdDev
            ws.cell(row=ri, column=11,
                    value=f'=$H{ri}-0.5*$G{ri}')

            # Alternating fill
            fill = ALT_FILL if ri % 2 == 0 else PatternFill()
            for ci in range(1, 12):
                ws.cell(row=ri, column=ci).fill = fill
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")

            ri += 1

    ws.freeze_panes = "A2"
    print(f"  Config: {ri-2} rows (all formulas)")

# ── Sheet 3/4: HistData (formula-driven) ──────────────────────────────────
def build_histdata(wb, unit_type, sheet_name):
    ws = wb.create_sheet(sheet_name)
    BLOCK_SIZE = N_BINS + 2  # 1 header row + 30 data rows + 1 gap

    # Config rows: LT = rows 2-13, MT = rows 14-25
    config_offset = 2 if unit_type == "LT" else 2 + len(MEAS_COLS)

    for idx, (raw_col, label, col_letter) in enumerate(MEAS_COLS):
        cfg_row = config_offset + idx
        block_start = idx * BLOCK_SIZE + 1

        # Header row
        headers = ["BinCenter", "BinLo", "BinHi", "TotalCount",
                    "Normal", "Warning(2-3s)", "Outlier(>3s)", "CurveY"]
        ws.cell(row=block_start, column=1, value=f"{label} ({unit_type})")
        ws.cell(row=block_start, column=1).font = Font(bold=True, size=11)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=block_start, column=ci + 0 if ci == 1 else ci, value=h)
            # Actually write headers in cols A-H
        # Rewrite headers cleanly
        for ci, h in enumerate(headers, 1):
            ws.cell(row=block_start, column=ci, value=h)
            ws.cell(row=block_start, column=ci).font = Font(bold=True, color="1F4E79")

        val_rng = f"RawData!${col_letter}$2:${col_letter}${LAST_ROW}"
        ut_rng  = f"RawData!$C$2:$C${LAST_ROW}"
        mean_ref  = f"Config!$F${cfg_row}"
        sd_ref    = f"Config!$G${cfg_row}"
        nall_ref  = f"Config!$D${cfg_row}"
        bw_ref    = f"Config!$J${cfg_row}"
        bs_ref    = f"Config!$K${cfg_row}"

        for b in range(N_BINS):
            r = block_start + 1 + b

            # A: BinCenter
            ws.cell(row=r, column=1,
                    value=f'={bs_ref}+{b}*{bw_ref}+{bw_ref}/2')

            # B: BinLo
            ws.cell(row=r, column=2,
                    value=f'=$A{r}-{bw_ref}/2')

            # C: BinHi
            ws.cell(row=r, column=3,
                    value=f'=$A{r}+{bw_ref}/2')

            # D: TotalCount = COUNTIFS
            ws.cell(row=r, column=4,
                    value=f'=COUNTIFS({ut_rng},"{unit_type}",{val_rng},">="&$B{r},{val_rng},"<"&$C{r},{val_rng},"<>0")')

            # E: Normal (within 2 sigma)
            ws.cell(row=r, column=5,
                    value=f'=IF({sd_ref}=0,0,IF(ABS(($A{r}-{mean_ref})/{sd_ref})<=2,$D{r},0))')

            # F: Warning (2-3 sigma)
            ws.cell(row=r, column=6,
                    value=f'=IF({sd_ref}=0,0,IF(AND(ABS(($A{r}-{mean_ref})/{sd_ref})>2,ABS(($A{r}-{mean_ref})/{sd_ref})<=3),$D{r},0))')

            # G: Outlier (>3 sigma)
            ws.cell(row=r, column=7,
                    value=f'=IF({sd_ref}=0,0,IF(ABS(($A{r}-{mean_ref})/{sd_ref})>3,$D{r},0))')

            # H: CurveY = NORM.DIST * n_all * binwidth
            ws.cell(row=r, column=8,
                    value=f'=IFERROR(NORM.DIST($A{r},{mean_ref},{sd_ref},FALSE)*{nall_ref}*{bw_ref},0)')

            # Alternating fill
            fill = ALT_FILL if b % 2 == 0 else PatternFill()
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = fill

    ws.column_dimensions["A"].width = 12
    for ci in range(2, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    print(f"  {sheet_name}: {len(MEAS_COLS)} measurement blocks x {N_BINS} bins (all formulas)")

# ── Sheet 5/6: BellCurves (charts referencing HistData) ──────────────────
def build_charts(wb, unit_type, chart_sheet_name, hist_sheet_name):
    ws = wb.create_sheet(chart_sheet_name)
    ws.sheet_view.showGridLines = False
    hist_ws = wb[hist_sheet_name]

    BLOCK_SIZE = N_BINS + 2
    charts_per_row = 3
    chart_w = 12  # cm
    chart_h = 9   # cm

    for idx, (_, label, _) in enumerate(MEAS_COLS):
        block_start = idx * BLOCK_SIZE + 1
        data_start  = block_start + 1
        data_end    = block_start + N_BINS

        cats_ref = Reference(hist_ws, min_col=1, min_row=data_start, max_row=data_end)

        # Stacked bar with 3 zone series
        bar = BarChart()
        bar.type     = "col"
        bar.grouping = "stacked"
        bar.overlap  = 100
        bar.gapWidth = 15
        bar.title    = f"{label} [{unit_type}]"
        bar.y_axis.title = "Count"
        bar.x_axis.numFmt = "0.0"
        bar.x_axis.tickLblSkip = 5

        zone_info = [
            (5, "4472C4", "Normal"),
            (6, "ED7D31", "Warning 2-3s"),
            (7, "C00000", "Outlier >3s"),
        ]
        for col_idx, color, name in zone_info:
            ref = Reference(hist_ws, min_col=col_idx, min_row=data_start, max_row=data_end)
            bar.add_data(ref)
            s = bar.series[-1]
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.solidFill = color

        bar.set_categories(cats_ref)

        # Normal curve line overlay
        line = LineChart()
        curve_ref = Reference(hist_ws, min_col=8, min_row=data_start, max_row=data_end)
        line.add_data(curve_ref)
        line.series[0].smooth = True
        line.series[0].graphicalProperties.line.solidFill = "FF0000"
        line.series[0].graphicalProperties.line.width = 25000
        line.series[0].marker.symbol = "none"

        bar += line
        bar.width  = chart_w
        bar.height = chart_h

        grid_row = idx // charts_per_row
        grid_col = idx % charts_per_row
        anchor_col = grid_col * 16 + 1
        anchor_row = grid_row * 18 + 1
        ws.add_chart(bar, get_column_letter(anchor_col) + str(anchor_row))

    print(f"  {chart_sheet_name}: {len(MEAS_COLS)} charts")

# ── Sheet 7: Tolerance_Bands (formula-driven) ──────────────────────────────
def build_tolerance(wb):
    ws = wb.create_sheet("Tolerance_Bands")
    headers = ["Unit", "Measurement", "n_pass", "Mean", "StdDev",
               "3s Low", "3s High", "P1 Low", "P99 High",
               "IQR Low", "IQR High", "Rec. Low", "Rec. High",
               "Spread Ratio", "Agreement"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["O"].width = 18

    ri = 2
    config_offset_lt = 2
    config_offset_mt = 2 + len(MEAS_COLS)

    for ut in ["LT", "MT"]:
        cfg_base = config_offset_lt if ut == "LT" else config_offset_mt
        for idx, (raw_col, label, col_letter) in enumerate(MEAS_COLS):
            cfg_row = cfg_base + idx
            val_rng = f"RawData!${col_letter}$2:${col_letter}${LAST_ROW}"
            ut_rng  = f"RawData!$C$2:$C${LAST_ROW}"
            pf_rng  = f"RawData!$F$2:$F${LAST_ROW}"

            # A: Unit, B: Measurement
            ws.cell(row=ri, column=1, value=ut)
            ws.cell(row=ri, column=2, value=label)

            # C: n_pass = Config ref
            ws.cell(row=ri, column=3, value=f'=Config!$E${cfg_row}')

            # D: Mean = Config ref
            ws.cell(row=ri, column=4, value=f'=Config!$F${cfg_row}')

            # E: StdDev = Config ref
            ws.cell(row=ri, column=5, value=f'=Config!$G${cfg_row}')

            # F: 3sigma Low
            ws.cell(row=ri, column=6, value=f'=Config!$F${cfg_row}-3*Config!$G${cfg_row}')

            # G: 3sigma High
            ws.cell(row=ri, column=7, value=f'=Config!$F${cfg_row}+3*Config!$G${cfg_row}')

            # H: P1 Low — AGGREGATE(16,6,arr,percentile) = PERCENTILE.INC ignoring errors
            agg_tol_cond = f'(({ut_rng}="{ut}")*({pf_rng}="Pass")*({val_rng}<>0))'
            agg_tol_arr  = f'{val_rng}/{agg_tol_cond}'
            ws.cell(row=ri, column=8,
                    value=f'=IFERROR(AGGREGATE(16,6,{agg_tol_arr},0.01),"")')

            # I: P99 High
            ws.cell(row=ri, column=9,
                    value=f'=IFERROR(AGGREGATE(16,6,{agg_tol_arr},0.99),"")')

            # J: IQR Low = Q1 - 1.5*IQR
            # AGGREGATE(17,6,arr,q) = QUARTILE.INC ignoring errors
            ws.cell(row=ri, column=10,
                    value=f'=IFERROR(AGGREGATE(17,6,{agg_tol_arr},1)-1.5*(AGGREGATE(17,6,{agg_tol_arr},3)-AGGREGATE(17,6,{agg_tol_arr},1)),"")')

            # K: IQR High = Q3 + 1.5*IQR
            ws.cell(row=ri, column=11,
                    value=f'=IFERROR(AGGREGATE(17,6,{agg_tol_arr},3)+1.5*(AGGREGATE(17,6,{agg_tol_arr},3)-AGGREGATE(17,6,{agg_tol_arr},1)),"")')

            # L: Recommended Low = MAX(3sig, P1, IQR)
            ws.cell(row=ri, column=12,
                    value=f'=IFERROR(MAX($F{ri},$H{ri},$J{ri}),"")')

            # M: Recommended High = MIN(3sig, P99, IQR)
            ws.cell(row=ri, column=13,
                    value=f'=IFERROR(MIN($G{ri},$I{ri},$K{ri}),"")')

            # N: Spread Ratio = max_width / min_width
            ws.cell(row=ri, column=14,
                    value=f'=IFERROR(MAX($G{ri}-$F{ri},$I{ri}-$H{ri},$K{ri}-$J{ri})/MIN($G{ri}-$F{ri},$I{ri}-$H{ri},$K{ri}-$J{ri}),"")')

            # O: Agreement
            ws.cell(row=ri, column=15,
                    value=f'=IF($N{ri}="","",IF($N{ri}<1.2,"Good",IF($N{ri}<1.5,"Review","Poor - non-normal")))')

            # Formatting
            fill = ALT_FILL if ri % 2 == 0 else PatternFill()
            for ci in range(1, 16):
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
                ws.cell(row=ri, column=ci).fill = fill

            # Color recommended columns green
            ws.cell(row=ri, column=12).fill = GRN_FILL
            ws.cell(row=ri, column=13).fill = GRN_FILL

            ri += 1

    # Conditional formatting on Agreement column
    ws.conditional_formatting.add(
        f"O2:O{ri-1}",
        CellIsRule(operator="equal", formula=['"Good"'], fill=GRN_FILL))
    ws.conditional_formatting.add(
        f"O2:O{ri-1}",
        CellIsRule(operator="equal", formula=['"Review"'], fill=YLW_FILL))
    ws.conditional_formatting.add(
        f"O2:O{ri-1}",
        CellIsRule(operator="equal", formula=['"Poor - non-normal"'], fill=RED_FILL))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{ri-1}"
    print(f"  Tolerance_Bands: {ri-2} rows (all formulas)")

# ── Sheet 8: Outliers (Python convenience view) ───────────────────────────
def build_outliers(wb, df, meas_present):
    """Pre-filter rows where z-score > 3 (computed from pass-only stats via pandas).
    This is a convenience snapshot; user can also filter RawData Z columns."""
    ws = wb.create_sheet("Outliers")
    headers = ["SerialNo", "UnitType", "BomNo", "TestDate", "OverallPassFail",
               "Failed?", "Measurement", "Value", "Z_Score"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Compute pass-only stats in pandas for outlier detection
    rows = []
    for raw_col, label, _ in MEAS_COLS:
        if raw_col not in meas_present:
            continue
        for ut in ["LT", "MT"]:
            mask_ut   = df["UnitType"] == ut
            mask_pass = mask_ut & (df["OverallPassFail"].str.strip().str.lower() == "pass")
            s_pass = pd.to_numeric(df.loc[mask_pass, raw_col], errors="coerce")
            s_pass = s_pass[s_pass.notna() & (s_pass != 0)]
            if len(s_pass) < 5:
                continue
            mu, sd = s_pass.mean(), s_pass.std()
            if sd == 0:
                continue
            s_all = pd.to_numeric(df.loc[mask_ut, raw_col], errors="coerce")
            z = (s_all - mu) / sd
            outlier_idx = mask_ut & (np.abs(z) > 3) & s_all.notna() & (s_all != 0)
            for i in df[outlier_idx].index:
                rows.append({
                    "SerialNo":        df.at[i, "SerialNo"],
                    "UnitType":        ut,
                    "BomNo":           df.at[i, "BomNo"],
                    "TestDate":        df.at[i, "TestDate"],
                    "OverallPassFail": df.at[i, "OverallPassFail"],
                    "Failed?":         "YES" if str(df.at[i, "OverallPassFail"]).lower() == "fail" else "no",
                    "Measurement":     label,
                    "Value":           round(float(s_all.at[i]), 4),
                    "Z_Score":         round(abs(float(z.at[i])), 2),
                })

    rows.sort(key=lambda r: r["Z_Score"], reverse=True)
    for ri, row in enumerate(rows, 2):
        alt = ALT_FILL if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row[key])
            c.alignment = Alignment(horizontal="center")
            if key == "Failed?" and row[key] == "YES":
                c.fill = RED_FILL
                c.font = Font(bold=True)
            else:
                c.fill = alt

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows)+1}"
    print(f"  Outliers: {len(rows)} records (snapshot)")

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    print("Loading data...")
    df, meas_present = load()

    wb = Workbook()
    wb.remove(wb.active)
    # Force Excel to recalculate all formulas when file is opened
    wb.calculation.fullCalcOnLoad = True

    print("Building sheets...")
    build_rawdata(wb, df, meas_present)
    build_config(wb)
    build_histdata(wb, "LT", "LT_HistData")
    build_histdata(wb, "MT", "MT_HistData")
    build_charts(wb, "LT", "LT_BellCurves", "LT_HistData")
    build_charts(wb, "MT", "MT_BellCurves", "MT_HistData")
    build_tolerance(wb)
    build_outliers(wb, df, meas_present)

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")

if __name__ == "__main__":
    main()
