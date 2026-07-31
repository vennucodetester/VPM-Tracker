import os
import tempfile
import unittest

from openpyxl import load_workbook

from models.task_node import TaskNode
from utils.excel_export import export_projects


def _node(name, potential=None, realized=None, status="In Progress"):
    node = TaskNode(name)
    node.vave_potential = potential
    node.vave_realized = realized
    node.status = status
    node.end_date = "2026-09-15"
    return node


def _attach(parent, children):
    parent.children = list(children)
    for child in parent.children:
        child.parent = parent
    return parent


class VaveSlideExportTests(unittest.TestCase):
    def _export(self, projects):
        tmp = tempfile.TemporaryDirectory()
        filename = os.path.join(tmp.name, "export.xlsx")
        export_projects(projects, filename)
        return tmp, load_workbook(filename, data_only=False)

    def test_non_vave_export_keeps_original_two_sheet_shape(self):
        project = {
            "name": "Standard",
            "metadata": {},
            "roots": [_node("Ordinary task")],
            "is_vave": False,
        }
        tmp, wb = self._export([project])
        self.addCleanup(tmp.cleanup)

        self.assertEqual(wb.sheetnames, ["Standard Tasks", "Standard Metadata"])
        headers = [cell.value for cell in wb["Standard Tasks"][1]]
        self.assertNotIn("Potential $", headers)
        self.assertNotIn("Realized $", headers)

    def test_vave_export_creates_data_and_slide_ready_previews(self):
        cassette = _attach(
            _node("Cassette VAVE activities"),
            [_node(f"Cassette idea {i}", potential=float(i)) for i in range(1, 9)],
        )
        case = _attach(
            _node("Case VAVE activities"),
            [_node(f"Case idea {i}", potential=float(i)) for i in range(1, 8)],
        )
        activity_root = _attach(_node("VAVE activities"), [cassette, case])
        project = {
            "name": "VAVE Project",
            "metadata": {},
            "roots": [activity_root],
            "is_vave": True,
        }
        tmp, wb = self._export([project])
        self.addCleanup(tmp.cleanup)

        self.assertIn("VAVE Project VAVE Data", wb.sheetnames)
        self.assertIn("Cassette VAVE activities", wb.sheetnames)
        self.assertIn("Case VAVE activities", wb.sheetnames)

        data = wb["VAVE Project VAVE Data"]
        self.assertEqual(data.max_row - 1, 15)
        self.assertEqual(data["N2"].value, "On Track")
        cassette_positions = [
            (data.cell(row, 5).value, data.cell(row, 6).value)
            for row in range(2, 10)
        ]
        self.assertEqual(cassette_positions[:4],
                         [("Left", 1), ("Left", 2), ("Left", 3), ("Left", 4)])
        self.assertEqual(cassette_positions[4:],
                         [("Right", 1), ("Right", 2), ("Right", 3), ("Right", 4)])

        cassette_preview = wb["Cassette VAVE activities"]
        case_preview = wb["Case VAVE activities"]
        self.assertEqual(cassette_preview["D6"].value, "On Track")
        self.assertEqual(case_preview["I6"].value, "On Track")
        self.assertEqual(
            cassette_preview["I3"].value,
            '=SUMIFS(\'VAVE Project VAVE Data\'!$K$2:$K$16,'
            '\'VAVE Project VAVE Data\'!$C$2:$C$16,"Cassette VAVE activities",'
            '\'VAVE Project VAVE Data\'!$O$2:$O$16,1)',
        )
        self.assertEqual(
            case_preview["I3"].value,
            '=SUMIFS(\'VAVE Project VAVE Data\'!$K$2:$K$16,'
            '\'VAVE Project VAVE Data\'!$C$2:$C$16,"Case VAVE activities",'
            '\'VAVE Project VAVE Data\'!$O$2:$O$16,1)',
        )

    def test_explicit_realized_value_controls_slide_savings_and_status(self):
        section = _attach(
            _node("Savings VAVE activities"),
            [_node("Implemented idea", potential=25.0, realized=18.0,
                   status="Completed")],
        )
        project = {
            "name": "VAVE",
            "metadata": {},
            "roots": [section],
            "is_vave": True,
        }
        tmp, wb = self._export([project])
        self.addCleanup(tmp.cleanup)

        data = wb["VAVE VAVE Data"]
        self.assertEqual(data["K2"].value, 18.0)
        self.assertEqual(data["L2"].value, "Realized")
        self.assertEqual(data["N2"].value, "REALIZED")
        preview = wb["Savings VAVE activities"]
        self.assertEqual(preview["G6"].value, 18.0)
        self.assertEqual(preview["I6"].value, "REALIZED")

    def test_more_than_fourteen_rows_creates_continuation_sheet(self):
        section = _attach(
            _node("Large VAVE activities"),
            [_node(f"Idea {i}", potential=float(i)) for i in range(1, 16)],
        )
        project = {
            "name": "VAVE",
            "metadata": {},
            "roots": [section],
            "is_vave": True,
        }
        tmp, wb = self._export([project])
        self.addCleanup(tmp.cleanup)

        self.assertIn("Large VAVE activities", wb.sheetnames)
        self.assertIn("Large VAVE activities (2)", wb.sheetnames)
        self.assertEqual(wb["Large VAVE activities (2)"]["A1"].value,
                         "Large VAVE activities (continued)")
        data = wb["VAVE VAVE Data"]
        self.assertEqual(data["D16"].value, 2)
        self.assertEqual(data["E16"].value, "Full")
        self.assertEqual(data["F16"].value, 1)


if __name__ == "__main__":
    unittest.main()
