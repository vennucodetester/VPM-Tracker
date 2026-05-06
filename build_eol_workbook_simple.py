"""
build_eol_workbook_simple.py  v4 — 3-Sigma Only, Simplified
Measurements: Compressor Discharge, Coil Outlet, Air Discharge,
              Ambient Temp, Humidity, Startup Voltage, Final Voltage,
              Startup Current, Final Current.
Tolerance method: pass-only Mean ± 3σ exclusively.
Output: EOL_BellCurves_v4_3sigma.xlsx
"""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
BASE      = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE, "SAVE FILES", "Run test data start to 4-27-26.xlsx")
OUT_DIR   = os.path.join(BASE, "eol_output")
OUT_FILE  = os.path.join(OUT_DIR, "EOL_BellCurves_v4_3sigma.xlsx")

BOM_MAP = {"CRS0000675": "LT", "CRS0000674": "MT"}

# (raw_col_name, display_label, RawData_column_letter)
MEAS_COLS = [
    ("RunTestSuctionTubeTemp",       "Compressor Discharge",  "G"),
    ("RunTestEvapInletRetBendATemp", "Coil Outlet",           "H"),
    ("RunTestEvapOutletAirTemp",     "Air Discharge",         "I"),
    ("RunTestAmbientTemperature",    "Ambient Temp",          "J"),
    ("RunTestRelativeHumidity",      "Humidity",              "K"),
    ("RunTestVoltageStartup",        "Startup Voltage",       "L"),
    ("RunTestFinalVoltage",          "Final Voltage",         "M"),
    ("RunTestCurrentStartup",        "Startup Current",       "N"),
    ("RunTestFinalCurrent",          "Final Current",         "O"),
]
N_MEAS   = len(MEAS_COLS)
N_BINS   = 30
LAST_ROW = 941   # row 1 = header, rows 2-941 = 940 data rows

# ── Styles ─────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL  = PatternFill("solid", fgColor="EEF4FB")   # light blue
RED_FILL  = PatternFill("solid", fgColor="FF9999")
GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
ORG_FILL  = PatternFill("solid", fgColor="FCE4D6")   # light orange for LT
BLU_FILL  = PatternFill("solid", fgColor="DDEEFF")   # light blue for MT

THIN = Side(style="thin", color="B0B0B0")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_cell(ws, row, col, val, width=16):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col)].width = width
    return c

def set_center(ws, row, col, val=None, fill=None, bold=False, size=10, num_fmt=None):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        c.fill = fill
    if bold:
        c.font = Font(bold=True, size=size)
    else:
        c.font = Font(size=size)
    if num_fmt:
        c.number_format = num_fmt
    c.border = BOX
    return c

# ── Helper: which _Helpers column for a given measurement + variant ────────
def helper_col(meas_idx, variant):
    """variant: 'LT_all', 'MT_all', 'LT_pass', 'MT_pass' → column letter"""
    offset = {"LT_all": 1, "MT_all": 2, "LT_pass": 3, "MT_pass": 4}[variant]
    return get_column_letter(4 * meas_idx + offset)

# ── Load data ───────────────────────────────────────────────────────────────
def load():
    df = pd.read_excel(DATA_FILE, engine="openpyxl")
    df["UnitType"] = df["BomNo"].map(BOM_MAP)
    df["TestDate"] = pd.to_datetime(df["TestDate"]).dt.date
    id_cols = ["SerialNo", "BomNo", "UnitType", "ModelNo", "TestDate", "OverallPassFail"]
    meas_present = [c for c, _, _ in MEAS_COLS if c in df.columns]
    missing = [c for c, _, _ in MEAS_COLS if c not in df.columns]
    if missing:
        print(f"  WARNING: columns not found in source data: {missing}")
    return df[id_cols + meas_present].copy(), meas_present

# ── Sheet 1: RawData ────────────────────────────────────────────────────────
def build_rawdata(wb, df, meas_present):
    ws = wb.create_sheet("RawData")
    id_cols = ["SerialNo", "BomNo", "UnitType", "ModelNo", "TestDate", "OverallPassFail"]

    # Build display labels map
    label_map = {raw: lbl for raw, lbl, _ in MEAS_COLS}
    z_headers  = [f"Z_{label_map.get(c, c)}" for c in meas_present]
    disp_meas  = [label_map.get(c, c) for c in meas_present]
    headers    = id_cols + disp_meas + z_headers

    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h, width=max(14, len(h) * 0.9))

    n_id   = len(id_cols)
    n_meas = len(meas_present)
    z_col_start = n_id + n_meas + 1

    # Write source data
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(id_cols, 1):
            ws.cell(row=ri, column=ci, value=row[col])
        for mi, col in enumerate(meas_present):
            ws.cell(row=ri, column=n_id + mi + 1, value=row.get(col))

    # Z-score formulas referencing Config sheet
    # Config layout: LT rows 2-(N_MEAS+1), MT rows (N_MEAS+2)-(2*N_MEAS+1)
    for ri in range(2, LAST_ROW + 1):
        for mi in range(n_meas):
            val_cell = f"{get_column_letter(n_id + mi + 1)}{ri}"
            ut_cell  = f"$C{ri}"
            cfg_row_lt = mi + 2
            cfg_row_mt = mi + 2 + N_MEAS
            mean_ref = f'IF({ut_cell}="LT",Config!$F${cfg_row_lt},Config!$F${cfg_row_mt})'
            sd_ref   = f'IF({ut_cell}="LT",Config!$G${cfg_row_lt},Config!$G${cfg_row_mt})'
            formula  = f'=IF(OR({val_cell}=0,{val_cell}=""),"",ABS(({val_cell}-{mean_ref})/{sd_ref}))'
            ws.cell(row=ri, column=z_col_start + mi, value=formula)

    # Conditional formatting: red on z > 3
    for mi in range(n_meas):
        cl = get_column_letter(z_col_start + mi)
        ws.conditional_formatting.add(
            f"{cl}2:{cl}{LAST_ROW}",
            CellIsRule(operator="greaterThan", formula=["3"], fill=RED_FILL))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{LAST_ROW}"
    print(f"  RawData: {LAST_ROW-1} rows, {len(headers)} cols")

# ── Sheet: _Helpers ────────────────────────────────────────────────────────
def build_helpers(wb):
    ws = wb.create_sheet("_Helpers")
    ws.sheet_state = "hidden"

    for m_idx, (_, label, col_letter) in enumerate(MEAS_COLS):
        for v_idx, variant in enumerate(["LT_all", "MT_all", "LT_pass", "MT_pass"]):
            col = 4 * m_idx + v_idx + 1
            ws.cell(row=1, column=col, value=f"{label}_{variant}").font = Font(bold=True, size=8)
            ws.column_dimensions[get_column_letter(col)].width = 12

    for r in range(2, LAST_ROW + 1):
        for m_idx, (_, _, col_letter) in enumerate(MEAS_COLS):
            val_cell = f"RawData!${col_letter}{r}"
            ut_cell  = f"RawData!$C{r}"
            pf_cell  = f"RawData!$F{r}"
            ws.cell(row=r, column=4*m_idx+1,
                    value=f'=IF(AND({ut_cell}="LT",{val_cell}<>0),{val_cell},"")')
            ws.cell(row=r, column=4*m_idx+2,
                    value=f'=IF(AND({ut_cell}="MT",{val_cell}<>0),{val_cell},"")')
            ws.cell(row=r, column=4*m_idx+3,
                    value=f'=IF(AND({ut_cell}="LT",{pf_cell}="Pass",{val_cell}<>0),{val_cell},"")')
            ws.cell(row=r, column=4*m_idx+4,
                    value=f'=IF(AND({ut_cell}="MT",{pf_cell}="Pass",{val_cell}<>0),{val_cell},"")')

    print(f"  _Helpers: {LAST_ROW-1} rows x {4*N_MEAS} cols (hidden)")

# ── Sheet 2: Config ─────────────────────────────────────────────────────────
def build_config(wb):
    ws = wb.create_sheet("Config")
    headers = ["Unit", "Measurement", "RawDataCol", "n_all", "n_pass",
               "Mean_pass", "StdDev_pass", "Min_all", "Max_all", "BinWidth", "BinStart"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)

    ri = 2
    for ut in ["LT", "MT"]:
        for m_idx, (_, label, col_letter) in enumerate(MEAS_COLS):
            ws.cell(row=ri, column=1, value=ut)
            ws.cell(row=ri, column=2, value=label)
            ws.cell(row=ri, column=3, value=col_letter)

            all_col  = helper_col(m_idx, f"{ut}_all")
            pass_col = helper_col(m_idx, f"{ut}_pass")
            all_rng  = f"_Helpers!${all_col}$2:${all_col}${LAST_ROW}"
            pass_rng = f"_Helpers!${pass_col}$2:${pass_col}${LAST_ROW}"

            ws.cell(row=ri, column=4,  value=f'=COUNT({all_rng})')
            ws.cell(row=ri, column=5,  value=f'=COUNT({pass_rng})')
            ws.cell(row=ri, column=6,  value=f'=IFERROR(AVERAGE({pass_rng}),0)')
            ws.cell(row=ri, column=7,  value=f'=IFERROR(STDEV({pass_rng}),0)')
            ws.cell(row=ri, column=8,  value=f'=IFERROR(MIN({all_rng}),0)')
            ws.cell(row=ri, column=9,  value=f'=IFERROR(MAX({all_rng}),0)')
            ws.cell(row=ri, column=10, value=f'=IFERROR(($I{ri}-$H{ri}+$G{ri})/{N_BINS},1)')
            ws.cell(row=ri, column=11, value=f'=$H{ri}-0.5*$G{ri}')

            fill = ALT_FILL if ri % 2 == 0 else PatternFill()
            for ci in range(1, 12):
                ws.cell(row=ri, column=ci).fill = fill
                ws.cell(row=ri, column=ci).alignment = Alignment(horizontal="center")
            ri += 1

    ws.freeze_panes = "A2"
    print(f"  Config: {ri-2} rows")

# ── Sheets 3/4: HistData ───────────────────────────────────────────────────
def build_histdata(wb, unit_type, sheet_name):
    ws = wb.create_sheet(sheet_name)
    BLOCK_SIZE   = N_BINS + 2
    config_offset = 2 if unit_type == "LT" else 2 + N_MEAS

    for idx, (_, label, col_letter) in enumerate(MEAS_COLS):
        cfg_row     = config_offset + idx
        block_start = idx * BLOCK_SIZE + 1

        for ci, h in enumerate(["BinCenter","BinLo","BinHi","TotalCount",
                                  "Normal","Warning(2-3s)","Outlier(>3s)","CurveY"], 1):
            ws.cell(row=block_start, column=ci, value=h)
            ws.cell(row=block_start, column=ci).font = Font(bold=True, color="1F4E79")
        ws.cell(row=block_start, column=1, value=f"{label} ({unit_type})")
        ws.cell(row=block_start, column=1).font = Font(bold=True, size=11)

        val_rng  = f"RawData!${col_letter}$2:${col_letter}${LAST_ROW}"
        ut_rng   = f"RawData!$C$2:$C${LAST_ROW}"
        mean_ref = f"Config!$F${cfg_row}"
        sd_ref   = f"Config!$G${cfg_row}"
        nall_ref = f"Config!$D${cfg_row}"
        bw_ref   = f"Config!$J${cfg_row}"
        bs_ref   = f"Config!$K${cfg_row}"

        for b in range(N_BINS):
            r = block_start + 1 + b
            ws.cell(row=r, column=1, value=f'={bs_ref}+{b}*{bw_ref}+{bw_ref}/2')
            ws.cell(row=r, column=2, value=f'=$A{r}-{bw_ref}/2')
            ws.cell(row=r, column=3, value=f'=$A{r}+{bw_ref}/2')
            ws.cell(row=r, column=4,
                    value=f'=COUNTIFS({ut_rng},"{unit_type}",{val_rng},">="&$B{r},{val_rng},"<"&$C{r},{val_rng},"<>0")')
            ws.cell(row=r, column=5,
                    value=f'=IF({sd_ref}=0,0,IF(ABS(($A{r}-{mean_ref})/{sd_ref})<=2,$D{r},0))')
            ws.cell(row=r, column=6,
                    value=f'=IF({sd_ref}=0,0,IF(AND(ABS(($A{r}-{mean_ref})/{sd_ref})>2,ABS(($A{r}-{mean_ref})/{sd_ref})<=3),$D{r},0))')
            ws.cell(row=r, column=7,
                    value=f'=IF({sd_ref}=0,0,IF(ABS(($A{r}-{mean_ref})/{sd_ref})>3,$D{r},0))')
            ws.cell(row=r, column=8,
                    value=f'=IFERROR(NORMDIST($A{r},{mean_ref},{sd_ref},FALSE)*{nall_ref}*{bw_ref},0)')

            fill = ALT_FILL if b % 2 == 0 else PatternFill()
            for ci in range(1, 9):
                ws.cell(row=r, column=ci).fill = fill

    for ci in range(1, 9):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    print(f"  {sheet_name}: {N_MEAS} blocks x {N_BINS} bins")

# ── Sheets 5/6: Bell Curve Charts ─────────────────────────────────────────
def build_charts(wb, unit_type, chart_sheet_name, hist_sheet_name):
    ws      = wb.create_sheet(chart_sheet_name)
    ws.sheet_view.showGridLines = False
    hist_ws = wb[hist_sheet_name]

    BLOCK_SIZE     = N_BINS + 2
    charts_per_row = 3

    for idx, (_, label, _) in enumerate(MEAS_COLS):
        block_start = idx * BLOCK_SIZE + 1
        data_start  = block_start + 1
        data_end    = block_start + N_BINS

        cats_ref = Reference(hist_ws, min_col=1, min_row=data_start, max_row=data_end)

        bar = BarChart()
        bar.type     = "col"
        bar.grouping = "stacked"
        bar.overlap  = 100
        bar.gapWidth = 15
        bar.title    = f"{label} [{unit_type}]"
        bar.y_axis.title = "Count"
        bar.x_axis.numFmt      = "0.0"
        bar.x_axis.tickLblSkip = 5

        for col_idx, color, name in [(5, "4472C4", "Normal"),
                                      (6, "ED7D31", "Warning 2-3σ"),
                                      (7, "C00000", "Outlier >3σ")]:
            ref = Reference(hist_ws, min_col=col_idx, min_row=data_start, max_row=data_end)
            bar.add_data(ref)
            s = bar.series[-1]
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.solidFill = color

        bar.set_categories(cats_ref)

        line = LineChart()
        curve_ref = Reference(hist_ws, min_col=8, min_row=data_start, max_row=data_end)
        line.add_data(curve_ref)
        line.series[0].smooth = True
        line.series[0].graphicalProperties.line.solidFill = "FF0000"
        line.series[0].graphicalProperties.line.width     = 25000
        line.series[0].marker.symbol = "none"

        bar += line
        bar.width  = 12
        bar.height = 9

        grid_row   = idx // charts_per_row
        grid_col   = idx % charts_per_row
        anchor_col = grid_col * 16 + 1
        anchor_row = grid_row * 18 + 1
        ws.add_chart(bar, get_column_letter(anchor_col) + str(anchor_row))

    print(f"  {chart_sheet_name}: {N_MEAS} charts")

# ── Sheet 7: Tolerance_Bands — 3-Sigma Only ────────────────────────────────
def build_tolerance(wb):
    ws = wb.create_sheet("Tolerance_Bands")
    headers = ["Unit", "Measurement", "n_pass", "Mean", "Std Dev",
               "3σ Low", "3σ High"]
    widths  = [8, 22, 10, 12, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        hdr_cell(ws, 1, ci, h, width=w)

    ws.row_dimensions[1].height = 30

    ri = 2
    for ut in ["LT", "MT"]:
        cfg_base = 2 if ut == "LT" else 2 + N_MEAS
        for idx, (_, label, _) in enumerate(MEAS_COLS):
            cfg_row = cfg_base + idx
            fill    = ORG_FILL if ut == "LT" else BLU_FILL

            ws.cell(row=ri, column=1, value=ut)
            ws.cell(row=ri, column=2, value=label)

            # n_pass, Mean, StdDev from Config
            ws.cell(row=ri, column=3, value=f'=Config!$E${cfg_row}')
            ws.cell(row=ri, column=4, value=f'=Config!$F${cfg_row}')
            ws.cell(row=ri, column=5, value=f'=Config!$G${cfg_row}')

            # 3-sigma Low and High
            ws.cell(row=ri, column=6, value=f'=Config!$F${cfg_row}-3*Config!$G${cfg_row}')
            ws.cell(row=ri, column=7, value=f'=Config!$F${cfg_row}+3*Config!$G${cfg_row}')

            for ci in range(1, 8):
                c = ws.cell(row=ri, column=ci)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = BOX
                c.fill      = fill
                if ci in (6, 7):
                    c.font      = Font(bold=True, size=10)
                    c.number_format = "0.00"
                else:
                    c.number_format = "0.00"

            ws.row_dimensions[ri].height = 20
            ri += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ri-1}"
    print(f"  Tolerance_Bands: {ri-2} rows (3-sigma only)")

# ── Sheet 8: Machine_Limits — clean programmer-ready summary ───────────────
def build_machine_limits(wb):
    ws = wb.create_sheet("Machine_Limits")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1,
                    value="EOL TEST MACHINE LIMITS  —  3-Sigma (Pass Units Only)")
    title.font      = Font(bold=True, size=14, color="FFFFFF")
    title.fill      = HDR_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-headers row
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:D2")
    ws.merge_cells("E2:G2")

    def sub(row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill      = fill or HDR_FILL
        c.border    = BOX
        return c

    sub(2, 1, "Measurement")
    sub(2, 2, "n\n(Pass Units)")
    sub(2, 3, "LT Unit (CRS0000675)", PatternFill("solid", fgColor="C65911"))
    sub(2, 5, "MT Unit (CRS0000674)", PatternFill("solid", fgColor="1F4E79"))

    for col, val in [(3, "3σ Low"), (4, "3σ High"), (5, "3σ Low"), (6, "3σ High"), (7, "Mean (LT / MT)")]:
        c = ws.cell(row=3, column=col, value=val)
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor="C65911") if col in (3, 4) else \
                      PatternFill("solid", fgColor="1F4E79") if col in (5, 6) else HDR_FILL
        c.border    = BOX

    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    for col in ["C","D","E","F"]:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["G"].width = 20

    # Config row layout: LT rows 2-(N_MEAS+1), MT rows (N_MEAS+2)-(2*N_MEAS+1)
    lt_base = 2
    mt_base = 2 + N_MEAS

    for idx, (_, label, _) in enumerate(MEAS_COLS):
        r       = idx + 4
        lt_row  = lt_base + idx
        mt_row  = mt_base + idx
        row_fill = ALT_FILL if idx % 2 == 0 else PatternFill()

        # Measurement name
        c = ws.cell(row=r, column=1, value=label)
        c.font      = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill      = row_fill
        c.border    = BOX

        # n_pass (LT / MT average shown as "LT / MT")
        c = ws.cell(row=r, column=2,
                    value=f'=Config!$E${lt_row}&" / "&Config!$E${mt_row}')
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = row_fill
        c.border    = BOX

        # LT 3σ Low
        c = ws.cell(row=r, column=3,
                    value=f'=Config!$F${lt_row}-3*Config!$G${lt_row}')
        c.number_format = "0.00"
        c.font      = Font(bold=True, size=10)
        c.fill      = ORG_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

        # LT 3σ High
        c = ws.cell(row=r, column=4,
                    value=f'=Config!$F${lt_row}+3*Config!$G${lt_row}')
        c.number_format = "0.00"
        c.font      = Font(bold=True, size=10)
        c.fill      = ORG_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

        # MT 3σ Low
        c = ws.cell(row=r, column=5,
                    value=f'=Config!$F${mt_row}-3*Config!$G${mt_row}')
        c.number_format = "0.00"
        c.font      = Font(bold=True, size=10)
        c.fill      = BLU_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

        # MT 3σ High
        c = ws.cell(row=r, column=6,
                    value=f'=Config!$F${mt_row}+3*Config!$G${mt_row}')
        c.number_format = "0.00"
        c.font      = Font(bold=True, size=10)
        c.fill      = BLU_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BOX

        # Mean (LT / MT) as text
        c = ws.cell(row=r, column=7,
                    value=f'=TEXT(Config!$F${lt_row},"0.00")&" / "&TEXT(Config!$F${mt_row},"0.00")')
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = row_fill
        c.border    = BOX

        ws.row_dimensions[r].height = 22

    # Footer note
    footer_row = N_MEAS + 4
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    fn = ws.cell(row=footer_row, column=1,
                 value="Limits derived from pass-only units. Mean ± 3σ captures 99.7% of normal production. "
                       "All values are live Excel formulas — click any cell to trace back to source data.")
    fn.font      = Font(italic=True, size=9, color="595959")
    fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 30

    print(f"  Machine_Limits: {N_MEAS} measurements, LT + MT side by side")

# ── Sheet 9: Outliers ──────────────────────────────────────────────────────
def build_outliers(wb, df, meas_present):
    ws = wb.create_sheet("Outliers")
    headers = ["SerialNo", "UnitType", "BomNo", "TestDate", "OverallPassFail",
               "Failed?", "Measurement", "Value", "Z_Score"]
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, 1, ci, h)

    label_map = {raw: lbl for raw, lbl, _ in MEAS_COLS}
    rows = []
    for raw_col, label, _ in MEAS_COLS:
        if raw_col not in meas_present:
            continue
        for ut in ["LT", "MT"]:
            mask_ut   = df["UnitType"] == ut
            mask_pass = mask_ut & (df["OverallPassFail"].str.strip().str.lower() == "pass")
            s_pass = pd.to_numeric(df.loc[mask_pass, raw_col], errors="coerce")
            s_pass = s_pass[s_pass.notna() & (s_pass != 0)]
            if len(s_pass) < 5:
                continue
            mu, sd = s_pass.mean(), s_pass.std()
            if sd == 0:
                continue
            s_all = pd.to_numeric(df.loc[mask_ut, raw_col], errors="coerce")
            z = (s_all - mu) / sd
            outlier_idx = mask_ut & (np.abs(z) > 3) & s_all.notna() & (s_all != 0)
            for i in df[outlier_idx].index:
                rows.append({
                    "SerialNo":        df.at[i, "SerialNo"],
                    "UnitType":        ut,
                    "BomNo":           df.at[i, "BomNo"],
                    "TestDate":        df.at[i, "TestDate"],
                    "OverallPassFail": df.at[i, "OverallPassFail"],
                    "Failed?":         "YES" if str(df.at[i, "OverallPassFail"]).lower() == "fail" else "no",
                    "Measurement":     label,
                    "Value":           round(float(s_all.at[i]), 4),
                    "Z_Score":         round(abs(float(z.at[i])), 2),
                })

    rows.sort(key=lambda r: r["Z_Score"], reverse=True)
    for ri, row in enumerate(rows, 2):
        alt = ALT_FILL if ri % 2 == 0 else PatternFill()
        for ci, key in enumerate(headers, 1):
            c = ws.cell(row=ri, column=ci, value=row[key])
            c.alignment = Alignment(horizontal="center")
            if key == "Failed?" and row[key] == "YES":
                c.fill = RED_FILL
                c.font = Font(bold=True)
            else:
                c.fill = alt

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(rows)+1}"
    print(f"  Outliers: {len(rows)} records")

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    print("Loading data...")
    df, meas_present = load()

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True

    print("Building sheets...")
    build_rawdata(wb, df, meas_present)
    build_helpers(wb)
    build_config(wb)
    build_histdata(wb, "LT", "LT_HistData")
    build_histdata(wb, "MT", "MT_HistData")
    build_charts(wb, "LT", "LT_BellCurves", "LT_HistData")
    build_charts(wb, "MT", "MT_BellCurves", "MT_HistData")
    build_tolerance(wb)
    build_machine_limits(wb)
    build_outliers(wb, df, meas_present)

    wb.save(OUT_FILE)
    print(f"\nSaved: {OUT_FILE}")

if __name__ == "__main__":
    main()
